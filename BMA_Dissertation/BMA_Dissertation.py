from asyncio.windows_events import NULL
from math import fabs
from typing import Text
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import argparse
import csv
import plotly.express as py
import numpy as np
import sys

def compute_bma(file_path, sheet_name, weight, ap, outputfile, imagefile):
    # Set up Plot Variables
    newarray = True
    newyap = True
    # Load Excel data
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    wb_obj = openpyxl.load_workbook(file_path)
    ds = wb_obj[sheet_name]
    maxrow = ds.max_row
    maxcol = ds.max_column
    # Validate required columns
    required = {weight, ap}
    if not required.issubset(df.columns):
        raise ValueError(f"Excel sheet must contain columns: " + str(required))

    headerrow = ds[1]
    #print(headerrow)
    j=1
    cellindex=0
    # Write to Excel output file
    with open(outputfile, 'w', newline='') as csvfile:
       fieldnames = ['Instance', 'BMA Weighted Prediction']
       writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
       writer.writeheader()
       for cell in headerrow:
          if cell.value != None:
             if weight == cell.value:
                #print(str(cell.value) + " " + str(cellindex))
                weightcol = cellindex
             if ap == cell.value:
                #print(str(cell.value) + " " + str(cellindex))
                apcol = cellindex
             if "nstance" in str(cell.value) : 
                posterior = 0
                total = 0
                isit = 0
                isitweight = 0
                isitSynthetic = 0
                bma_prediction1 = 0
                for index in range(1,maxcol+1):
                   #if (ds[index][j].value == 'A' or ds[index][j].value == 'x'):
                   if ds[index][j].value != None:
                       if (isinstance(ds[index][weightcol].value,float)) and (isinstance(ds[index][apcol].value,float)) :
                          if (isinstance(ds[index][j].value, float) and ds[index][j].value <= 1.000001):
                              isit = isit + (float(ds[index][weightcol].value) * float(ds[index][apcol].value) * float(ds[index][j].value))
                              isitweight = isitweight +  ds[index][weightcol].value
                          posterior = posterior + (float(ds[index][weightcol].value) * float(ds[index][apcol].value))
                          total = total + ds[index][weightcol].value
                if (total > 0):      
                   bma_prediction1=posterior/total
                   if (isit > 0):
                       isitSynthetic = isit/isitweight
                   writer.writerow({'Instance': cell.value, 'BMA Weighted Prediction': bma_prediction1})
                   #print(f"\n {cell.value} BMA Weighted Prediction: {bma_prediction1:.4f}")
                   if (newarray):
                       xinstance = np.array(cell.value)
                       yap = np.array(bma_prediction1)
                       if (isitSynthetic > 0):
                          isityap = np.array(isitSynthetic)
                          isitinstance = np.array(cell.value)
                          newyap=False
                       newarray = False;
                   else: 
                        xinstance = np.append(xinstance, np.array(cell.value))
                        yap = np.append(yap, np.array(bma_prediction1))
                        if (isitSynthetic > 0):
                           if (newyap):
                              isityap = np.array(isitSynthetic)
                              isitinstance = np.array(cell.value)
                              newyap = False
                           else: 
                               isityap = np.append(isityap, np.array(isitSynthetic))
                               isitinstance = np.append(isitinstance, np.array(cell.value))
                
             j = j+1
             cellindex=cellindex + 1

       # Compute unnormalized posterior
       df['UnnormalizedPosterior'] = df[weight] * df[ap]

       # Normalize to get posterior probabilities.  Assumes all values in the cell are floats
       total1 = df[weight].sum()
       df['Posterior'] = df['UnnormalizedPosterior'].sum() / total1

    # Compute BMA-weighted prediction
       bma_prediction = df['UnnormalizedPosterior'].sum() / total1

    # Display results
       xinstance = np.append(xinstance, np.array('All Method BMA Weighted Prediction'))
       #xinstance = np.array('All Method BMA Weighted Prediction')
       yap = np.append(yap, np.array(bma_prediction))
       #yap = np.array(bma_prediction)
       writer.writerow({'Instance': 'All Method BMA Weighted Prediction', 'BMA Weighted Prediction': bma_prediction})
       #print(f"\nAll Method BMA Weighted Prediction: {bma_prediction:.4f}")

       #print(f"{xinstance}")
       #print(f"{yap}")

       #fig = py.scatter(y=[0.948857, 0.53400475, 0.6510365, 0.5505247], x=['Instance 1','Instance 2', 'instance 3', 'Instance 4'], title="BMA Scatter Plot")
       #fig.show()
       fig = py.scatter(x=xinstance, y=yap, title="BMA Scatter Plot " + imagefile)
       fig.show()
       fig.write_image(imagefile)
       if (newyap == False):
          fig2 = py.scatter(x=isitinstance, y=isityap, title="BMA Synthetic Estimation " + imagefile)
          fig2.show()
          fig2.write_image('isit-' + imagefile)


# Example usage
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='BMA calculations from Excel Spreadsheet')
    if (len(sys.argv) < 5):
        print(f"Error running Python Script\n")
        print(f"usage: python BMA_Dissertation.py file_path weight_column averagepercision_column sheet_name outputfile(.csv) imagefile(file format: eg: jpg)")
        sys.exit(0);

    parser.add_argument('file_arg', type=Text, help='Provide full path for the filename.')
    parser.add_argument('weighted_arg', type=Text, help='Header Value for the weight given to the method (assume this is in col B)')
    parser.add_argument('acuracypercent_arg', type=Text, help='Header Value for the accuracy percentage to the method (assume this is in col C)')
    parser.add_argument('sheetname_arg', type=Text, help='The name of the sheet in the file that contains the data')
    parser.add_argument('outputfile_arg', type=Text, help='The path and name of the csv output file.')
    parser.add_argument('imagefile_arg', type=Text, help='The path and name of the scatter plot.')
    args = parser.parse_args()

    file_path = args.file_arg
    weight = args.weighted_arg
    ap = args.acuracypercent_arg
    sheet_name = args.sheetname_arg
    outputfile = args.outputfile_arg
    imagefile = args.imagefile_arg

    print(f"{file_path} {weight} {ap} {sheet_name} {outputfile} {imagefile}")
    #file_path = 'd:\\rhugh\\Documents\\CTU_doctoral\\Dissertation\\RawData.xlsx'
    compute_bma(file_path, sheet_name, weight, ap, outputfile, imagefile)
